import tkinter as tk
from tkinter import ttk

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style

import serial_2 as sp1

from openpyxl import Workbook
import pandas as pd
import numpy as np
import sys
import time

LARGE_FONT = ("Verdana", 12)
TITLE_FONT = ("Times",16)
NORM_FONT = ("Helvetica", 10)
SMALL_FONT = ("Helvetica", 8)
style.use("ggplot")


def init_gui():
    app = psoc_Sensors()
    # ani_resp = animation.FuncAnimation(resp_graph, recordResp, interval=100)
    ani_uv = animation.FuncAnimation(uv_graph, recordUV, interval=100)
    # ani_temp = animation.FuncAnimation(temp_graph, recordTemp, interval=100)
    app.geometry("1500x1000")
    app.mainloop()


def popupmsg(msg):
    popup = tk.Tk()
    popup.wm_title("Pop-Up Message!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.grid(row=0, column=0)
    b1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    b1.grid(row=0, column=1)
    popup.mainloop()


def record_start_uv(value):
    global UV_Record
    global start_time
    if value == 0:
        UV_Record = 1
        start_time = time.time()


def record_stop_uv(value):
    global UV_Record
    global start_time
    global UV1_Data
    global UV2_Data
    global UV3_Data
    global UV4_Data
    global UV5_Data
    global seconds

    if value == 1:
        UV_Record = 0
        start_time = 0
    # Saving the Data to an Excel Spread Sheet
    save_uv_data()
    UV1_Data = []
    UV2_Data = []
    UV3_Data = []
    UV4_Data = []
    UV5_Data = []
    seconds = []

def record_start_resp(value):
    global Resp_Record
    global start_time
    if value == 0:
        Resp_Record = 1
        start_time = time.time()


def record_stop_resp(value):
    global Resp_Record
    global start_time
    global Resp_Data
    global seconds

    if value == 1:
        Resp_Record = 0
        start_time = 0
    # Saving the Data to an Excel Spread Sheet
    save_resp_data()
    Resp_Data = []
    seconds = []


def record_start_temp(value):
    global Temp_Record
    global start_time
    if value == 0:
        Temp_Record = 1
        start_time = time.time()



def record_stop_temp(value):
    global Temp_Record
    global start_time
    global TMP116_Data
    global LMT70_Data
    global seconds


    if value == 1:
        Temp_Record = 0
        start_time = 0
    # Saving the Data to an Excel Spread Sheet
    save_temp_data()
    LMT70_Data = []
    TMP116_Data = []
    seconds = []

def print_Name1():
        print("Carl!")

def save_temp_data():
    global seconds
    global TMP116_Data
    global LMT70_Data
    dest_filename = 'C:/Users/Intern/Documents/Data/test-temp.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample #1"
    ws.cell(row=1, column=1, value="Time (s)")
    ws.cell(row=1, column=2, value="LMT70 (C)")
    ws.cell(row=1, column=3, value="TMP116 (C)")
    wb.save(filename = dest_filename)
    for i in range(1, len(seconds)):
        ws.cell(row=i+1, column=1, value=seconds[i-1])
        ws.cell(row=i+1, column=2, value=LMT70_Data[i-1])
        ws.cell(row=i+1, column=3, value=TMP116_Data[i-1])

    wb.save(filename = dest_filename)

def save_uv_data():
    global seconds
    global UV1_Data
    global UV2_Data
    global UV3_Data
    global UV4_Data
    global UV5_Data

    dest_filename = 'C:/Users/Intern/Documents/Data/test-uv.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample #1"
    ws.cell(row=1, column=1, value="Time (s)")
    ws.cell(row=1, column=2, value="UV1")
    ws.cell(row=1, column=3, value="UV2")
    ws.cell(row=1, column=4, value="UV3")
    ws.cell(row=1, column=5, value="UV4")
    ws.cell(row=1, column=6, value="UV5")

    for i in range(1, len(seconds)):
        ws.cell(row=i+1, column=1, value=seconds[i-1])
        ws.cell(row=i+1, column=2, value=UV1_Data[i-1])
        ws.cell(row=i+1, column=3, value=UV2_Data[i-1])
        ws.cell(row=i+1, column=4, value=UV3_Data[i-1])
        ws.cell(row=i+1, column=5, value=UV4_Data[i-1])
        ws.cell(row=i+1, column=6, value=UV5_Data[i-1])

    wb.save(filename = dest_filename)


def save_resp_data():
    global seconds
    global Resp_Data
    dest_filename = 'C:/Users/Intern/Documents/Data/test-resp.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample #1"
    ws.cell(row=1, column=1, value="Time (s)")
    ws.cell(row=1, column=2, value="SG % Change")

    for i in range(1,len(seconds)):
        ws.cell(row=i+1, column=1, value=seconds[i-1])
        ws.cell(row=i+1, column=4, value=Resp_Data[i-1])
    wb.save(filename = dest_filename)


def update_Resp_Fig():
    global resp_subplot
    resp_subplot.clear()
    resp_subplot.set_title("Respiration Breathing")
    resp_subplot.set_ylabel("Percent Strain Change (%)")  # Set ylabels
    resp_subplot.set_xlabel("Time (s)")
    if seconds[-1] < 20:
        resp_subplot.set_xlim([0, 20])
    else:
        time_low = seconds[-1] - 20
        resp_subplot.set_xlim([time_low,seconds[-1]])
    resp_subplot.set_ylim([0, 100])
    resp_subplot.ticklabel_format(useOffset=False)

def update_UV_Fig():
    global uv_subplot
    uv_subplot.clear()
    uv_subplot.set_title("UV Power Reading")
    uv_subplot.set_ylabel("Power Density mW/cm^2")  # Set ylabels
    uv_subplot.set_xlabel("Time (s)")
    if seconds[-1] < 20:
        uv_subplot.set_xlim([0, 20])
    else:
        time_low = seconds[-1] - 20
        uv_subplot.set_xlim([time_low,seconds[-1]])
    uv_subplot.set_ylim([0, 10])
    uv_subplot.ticklabel_format(useOffset=False)

def update_Temp_Fig():
    global temp_subplot
    temp_subplot.clear()
    temp_subplot.set_title("Temperature Reading")
    temp_subplot.set_ylabel("Celsius (C)")  # Set ylabels
    temp_subplot.set_xlabel("Time (s)")
    if seconds[-1] < 20:
        temp_subplot.set_xlim([0, 20])
    else:
        time_low = seconds[-1] - 20
        temp_subplot.set_xlim([time_low,seconds[-1]])
    temp_subplot.set_ylim([20, 25])
    temp_subplot.ticklabel_format(useOffset=False)



def update_total_Fig():
    print('Hello')


def recordUV(i):
    # i variable is basically a self
    # start data collection
    global UV_Record
    global UV1_Data
    global UV2_Data
    global UV3_Data
    global UV4_Data
    global UV5_Data
    global seconds

    if (UV_Record == 1):
        while (sp1.wait() == 0):
            pass
        UV1_value = sp1.text_to_decimal()
        UV2_value = sp1.text_to_decimal()
        UV3_value = sp1.text_to_decimal()
        UV4_value = sp1.text_to_decimal()
        UV5_value = sp1.text_to_decimal()

        count = time.time() - start_time

        UV1_Data.append(UV1_value)
        UV2_Data.append(UV2_value)
        UV3_Data.append(UV3_value)
        UV4_Data.append(UV4_value)
        UV5_Data.append(UV5_value)
        seconds.append(count)

        update_UV_Fig()
        uv_subplot.plot(seconds, UV1_Data, seconds, UV2_Data, seconds, UV3_Data, seconds, UV4_Data, seconds, UV5_Data)

def recordTemp(i):
    # i variable is basically a self
    # start data collection
    global Temp_Record
    global TMP116_Data
    global LMT70_Data
    global seconds

    if (Temp_Record == 1):
        while (sp1.wait() == 0):
            pass
        lmt70_value = sp1.text_to_decimal()
        tmp116_value = sp1.text_to_decimal()

        count = time.time() - start_time

        LMT70_Data.append(lmt70_value)
        TMP116_Data.append(tmp116_value)
        seconds.append(count)

        update_Temp_Fig()
        temp_subplot.plot(seconds, LMT70_Data, seconds, TMP116_Data)


def recordResp(i):
    # i variable is basically a self
    # start data collection
    global Resp_Record
    global Resp_Data
    global seconds

    if (Resp_Record == 1):
        while (sp1.wait() == 0):
            pass
        Resp_value = sp1.read_dec()

        count = time.time() - start_time

        Resp_Data.append(Resp_value)
        seconds.append(count)

        update_Resp_Fig()
        resp_subplot.plot(seconds, Resp_Data)

class psoc_Sensors(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        #tk.Tk.iconbitmap(self, default="sw-caution.gif")
        tk.Tk.wm_title(self, "Wearable Sensor BLE Gui")

        container = tk.Frame(self)
        container.grid(row=0,column=0)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.create_menu(container)
        self.frames = {}

        for F in (Intro_Page, Help_Page, Main_Page):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(Intro_Page)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

    def create_menu(self, container):
        menubar = tk.Menu(container)

        fileMenu = tk.Menu(menubar, tearoff=0)
        fileMenu.add_command(label="Save Graph", command=lambda: popupmsg("Not supported just yet!"))
        fileMenu.add_command(label="Export Excel", command=lambda: popupmsg("Not supported just yet!"))
        fileMenu.add_separator()
        fileMenu.add_command(label="Exit", command=quit)
        menubar.add_cascade(label="File", menu=fileMenu)

        subMenu = tk.Menu(menubar, tearoff=1)
        subMenu.add_command(label="Sensor Calibration", command=lambda: popupmsg("Not supported just yet!"))
        subMenu.add_command(label="Save settings", command=lambda: popupmsg("Not supported just yet!"))
        menubar.add_cascade(label="Settings", menu=subMenu)

        editMenu = tk.Menu(menubar, tearoff=1)
        editMenu.add_command(label="Help Menu", command=lambda: self.show_frame(Help_Page))
        editMenu.add_separator()
        editMenu.add_command(label="Print Name", command=lambda: print_Name1())
        menubar.add_cascade(label="Help", menu=editMenu)

        tk.Tk.config(self, menu=menubar)


class Intro_Page(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        label1 = tk.Label(self, text=("Welcome to Wearable Sensor BLE Gui!"), font=LARGE_FONT)
        label1.grid(row=0,column=0)

        label2 = tk.Label(self, text=("Designed by the Biomedical Wearable Lab @ UCSD"), font=TITLE_FONT)
        label2.grid(row=1, column=0)

        button1 = ttk.Button(self, text="Start!", command=lambda: controller.show_frame(Main_Page))
        button1.grid(row=2, column=0)

        # button2 = ttk.Button(self, text="Quit", command=_quit())
        # button2.grid(row=3, column=0)

        button3 = ttk.Button(self, text="Help Me!", command=lambda: controller.show_frame(Help_Page))
        button3.grid(row=4, column=0)


class Help_Page(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)

        label = tk.Label(self, text=("Welcome to Help Menu!"), font=TITLE_FONT)
        label.grid(row=0,column=0)

        button1 = tk.Button(self, text="go Back to intro Page", command=lambda: controller.show_frame(Intro_Page))
        button1.grid(row=1,column=0)

        button2 = tk.Button(self, text="Go Back to main Page", command=lambda: controller.show_frame(Main_Page))
        button2.grid(row=2,column=0)

class Main_Page(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.create_Button()

        # canvas_frame_resp = tk.Frame(self)
        # canvas_resp = FigureCanvasTkAgg(resp_graph, canvas_frame_resp)
        # canvas_resp.show()
        # canvas_resp.get_tk_widget().pack()
        # canvas_frame_resp.grid(row=6, column=0, columnspan=5)
        #
        canvas_frame_uv = tk.Frame(self)
        canvas_uv = FigureCanvasTkAgg(uv_graph, canvas_frame_uv)
        canvas_uv.show()
        canvas_uv.get_tk_widget().pack()
        canvas_frame_uv.grid(row=6, column=10, columnspan=5)

        # canvas_frame_temp = tk.Frame(self)
        # canvas_temp = FigureCanvasTkAgg(temp_graph, canvas_frame_temp)
        # canvas_temp.show()
        # canvas_temp.get_tk_widget().pack()
        # canvas_frame_temp.grid(row=10, column=10, columnspan=5)

        # toolbar_frame_resp = tk.Frame(self)
        # toolbar_resp = NavigationToolbar2TkAgg(canvas_resp, toolbar_frame_resp)
        # toolbar_resp.update()
        # canvas_resp._tkcanvas.pack(side=tk.TOP)
        # toolbar_frame_resp.grid(row=7, column=0, columnspan=2)
        #
        toolbar_frame_uv = tk.Frame(self)
        toolbar_uv = NavigationToolbar2TkAgg(canvas_uv, toolbar_frame_uv)
        toolbar_uv.update()
        canvas_uv._tkcanvas.pack(side=tk.TOP)
        toolbar_frame_uv.grid(row=7, column=10, columnspan=2)

        # toolbar_frame_temp = tk.Frame(self)
        # toolbar_temp = NavigationToolbar2TkAgg(canvas_temp, toolbar_frame_temp)
        # toolbar_temp.update()
        # canvas_temp._tkcanvas.pack(side=tk.TOP)
        # toolbar_frame_temp.grid(row=7, column=10, columnspan=2)

    def create_Button(self):
        button_frame_uv = tk.Frame(self)
        button1 = ttk.Button(button_frame_uv, text="Start", command=lambda: record_start_uv(UV_Record))
        button1.grid(row=0,column=0)
        button2 = ttk.Button(button_frame_uv, text="Stop", command=lambda: record_stop_uv(UV_Record))
        button2.grid(row=0,column=1)
        button_frame_uv.grid(row=4,column=0)

        button_frame_resp = tk.Frame(self)
        button3 = ttk.Button(button_frame_resp, text="Start", command=lambda: record_start_resp(Resp_Record))
        button3.grid(row=0,column=0)
        button4 = ttk.Button(button_frame_resp, text="Stop", command=lambda: record_stop_resp(Resp_Record))
        button4.grid(row=0,column=1)
        button_frame_resp.grid(row=5,column=0)

        button_frame_temp = tk.Frame(self)
        button5 = ttk.Button(button_frame_temp, text="Start", command=lambda: record_start_temp(Temp_Record))
        button5.grid(row=0,column=0)
        button6 = ttk.Button(button_frame_temp, text="Stop", command=lambda: record_stop_temp(Temp_Record))
        button6.grid(row=0,column=1)
        button_frame_temp.grid(row=6,column=0)

def main():
    init_gui()

if __name__ == "__main__":
    main()
