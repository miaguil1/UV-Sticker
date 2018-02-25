import matplotlib.pyplot as plt
import tkinter as tk
from openpyxl import Workbook
from tkinter.filedialog import askopenfilename, askdirectory
from PIL import ImageTk, Image
import sys, os
from datetime import date, datetime, time
import pandas
import numpy
import scipy
import statistics

def resource_path(relative_path):
    try:
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    except:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def init_gui():
    root = tk.Tk()
    app = SgGui(root)
    root.mainloop()


class SgGui(tk.Frame):

    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        self.parent = parent

        self.resistors = [0] * 10
        self.cypress_parse_filename = '-'
        self.excel_destination = '-'

        self.uv_data1 = []
        self.uv_time_data1 = []
        self.uv_data2 = []
        self.uv_time_data2 = []

        self.temp1 = [0]*2
        self.temp_data1 = []
        self.temp_time_data1 = []

        self.temp2 = [0] * 2
        self.temp_data2 = []
        self.temp_time_data2 = []

        self.day = datetime.now().day
        self.month = datetime.now().month
        self.year = datetime.now().year

        self.combined_destination = '-'

        parent.minsize(width=750, height=750)
        parent.maxsize(width=1500, height=750)
        parent.title("Strain Gauge GUI")

        title_page = tk.Label(parent, text="Flexible Electronics GUI")
        title_page.grid(row=0, column=0)

        calculate_button = tk.Button(self.parent, text="Calculate Proper Gain", command=lambda:
                                     [self.update_resistors(), self.calculate_strain()])
        calculate_button.grid(row=1, column=0)

        close_button = tk.Button(self.parent, text="Close", command=parent.quit)
        close_button.grid(row=1, column=1)

        cypress_ble_button = tk.Button(self.parent, text="Upload Cypress App BLE Log", command=lambda:
                                       self.cypress_ble_pop_up())
        cypress_ble_button.grid(row=1, column=2)

        cypress_ble_uv1_button = tk.Button(self.parent, text="Upload CySmart BLE UV Log #1", command=lambda:
                                       self.cysmart_ble_uv_pop_up(1))
        cypress_ble_uv1_button.grid(row=1, column=3)

        cypress_ble_uv2_button = tk.Button(self.parent, text="Upload CySmart BLE UV Log #2", command=lambda:
                                       self.cysmart_ble_uv_pop_up(2))
        cypress_ble_uv2_button.grid(row=1, column=3)

        cypress_ble_temp1_button = tk.Button(self.parent, text="Upload CySmart BLE Temperature Log #1", command=lambda:
                                       self.cysmart_ble_temp_pop_up(1))
        cypress_ble_temp1_button.grid(row=1, column=5)

        cypress_ble_temp2_button = tk.Button(self.parent, text="Upload CySmart BLE Temperature Log #2", command=lambda:
                                       self.cysmart_ble_temp_pop_up(2))
        cypress_ble_temp2_button.grid(row=1, column=6)

        cypress_ble_temp_comp_button = tk.Button(self.parent, text="Compare Multiple Sensors", command=lambda:
                                       self.temp_compare())
        cypress_ble_temp_comp_button.grid(row=1, column=7)

        self.r1_label = tk.Label(self.parent, text="R1 value in ohm")
        self.r1_label.grid(row=2, column=0)
        self.er1 = tk.IntVar(self.parent)
        self.er1.set(self.resistors[0])
        self.er1_entry = tk.Entry(self.parent, textvariable=self.er1, width=10)
        self.er1_entry.grid(row=2, column=1)

        self.r2_label = tk.Label(self.parent, text="R2 value in ohm")
        self.r2_label.grid(row=3, column=0)
        self.er2 = tk.IntVar(self.parent, value=self.resistors[1])
        self.er2_entry = tk.Entry(self.parent, textvariable=self.er2, width=10)
        self.er2_entry.grid(row=3, column=1)

        self.r3_label = tk.Label(self.parent, text="R3 value in ohm")
        self.r3_label.grid(row=4, column=0)
        self.er3 = tk.IntVar(self.parent, value=self.resistors[2])
        self.er3_entry = tk.Entry(self.parent, textvariable=self.er3, width=10)
        self.er3_entry.grid(row=4, column=1)

        self.r4_label = tk.Label(self.parent, text="R4 value in ohm")
        self.r4_label.grid(row=5, column=0)
        self.er4 = tk.IntVar(self.parent, value=self.resistors[3])
        self.er4_entry = tk.Entry(self.parent, textvariable=self.er4, width=10)
        self.er4_entry.grid(row=5, column=1)

        self.r5_label = tk.Label(self.parent, text="R5 value in ohm")
        self.r5_label.grid(row=6, column=0)
        self.er5 = tk.IntVar(self.parent, value=self.resistors[4])
        self.er5_entry = tk.Entry(self.parent, textvariable=self.er5, width=10)
        self.er5_entry.grid(row=6, column=1)

        self.r6_label = tk.Label(self.parent, text="R6 value in ohm")
        self.r6_label.grid(row=7, column=0)
        self.er6 = tk.IntVar(self.parent, value=self.resistors[5])
        self.er6_entry = tk.Entry(self.parent, textvariable=self.er6, width=10)
        self.er6_entry.grid(row=7, column=1)

        self.rx_label = tk.Label(self.parent, text="Rx (ohm)")
        self.rx_label.grid(row=8, column=0)
        self.erx = tk.IntVar(self.parent, value=self.resistors[6])
        self.erx_entry = tk.Entry(self.parent, textvariable=self.erx, width=10)
        self.erx_entry.grid(row=8, column=1)

        self.rx_tol_label = tk.Label(self.parent, text="Rx Tolerance (ohm)")
        self.rx_tol_label.grid(row=8, column=2)
        self.erx_tol = tk.IntVar(self.parent, value=self.resistors[7])
        self.erx_tol_entry = tk.Entry(self.parent, textvariable=self.erx_tol, width=10)
        self.erx_tol_entry.grid(row=8, column=3)

        self.rg1_label = tk.Label(self.parent, text="RG1")
        self.rg1_label.grid(row=9, column=0)
        self.rg1 = tk.IntVar(self.parent, value=self.resistors[8])
        self.rg1_value = tk.Label(self.parent, textvariable=self.rg1)
        self.rg1_value.grid(row=9, column=1)

        self.rg2_label = tk.Label(self.parent, text="RG2")
        self.rg2_label.grid(row=10, column=0)
        self.rg2 = tk.IntVar(self.parent, value=self.resistors[9])
        self.rg2_value = tk.Label(self.parent, textvariable=self.rg2)
        self.rg2_value.grid(row=10, column=1)

        self.bg_raw = Image.open(resource_path("Background_Circuit.png"))
        self.bg_rs_img = self.bg_raw.resize((750, 375), Image.ANTIALIAS)
        self.bg_img = ImageTk.PhotoImage(self.bg_rs_img)
        self.bg_label = tk.Label(self.parent, image=self.bg_img)
        self.bg_label.grid(row=11, column=0, rowspan=2, columnspan=6)

        self.resistor_entries = [self.er1_entry, self.er2_entry, self.er3_entry, self.er4_entry, self.er5_entry,
                                 self.er6_entry, self.erx_entry, self.erx_tol_entry]
        self.resistor_var = [self.er1, self.er2, self.er3, self.er4, self.er5,
                             self.er6, self.erx, self.erx_tol, self.rg1, self.rg2]

    def update_resistors(self):
        for i in range(0, 8):
            self.resistors[i] = int(self.resistor_entries[i].get())
            self.resistor_var[i].set(self.resistors[i])
        print(self.resistors)

    def calculate_strain(self):
        rg1 = self.resistors[0]*5
        rg2 = self.resistors[0] * 5
        self.resistor_var[8].set(rg1)
        self.resistor_var[9].set(rg2)
        v_supply = 3.3
        r1 = 977
        r2 = 23
        r3 = 1000
        r4 = 1000
        rx = 1002

        r5 = 1000
        r6 = 1000

        rg1 = 3.3E6
        rg2 = 3.3E6

        v1 = v_supply * r3 / (r1 + r2 + r3)
        v2 = v_supply * r4 / (r4 + rx)

        vamp_minus = v1 * rg1 / (rg1 + r5)
        vamp_plus = v2 * rg2 / (rg2 + r6)

        v_diff = abs(vamp_plus - vamp_minus)

        v_out = v_diff * (1 + (rg1 / r6))

        print(v1)
        print(v2)
        print(v_diff)
        print(v_out)

    def ask_user_file(self, file_show):
        self.cypress_parse_filename = askopenfilename()
        file_show.delete(0, tk.END)
        file_show.insert(0, self.cypress_parse_filename)

    def ask_user_output_folder(self, folder_show):
        self.excel_destination = askdirectory()
        folder_show.delete(0, tk.END)
        folder_show.insert(0, self.excel_destination)

    def cypress_ble_pop_up(self):
        popup_cypress = tk.Toplevel()
        popup_cypress.grab_set()
        popup_cypress.geometry("1000x100")
        popup_cypress.wm_title("Cypress BLE Data Upload")
        label = tk.Label(popup_cypress, text="Upload Cypress BLE Data Log")
        label.grid(row=0, column=0)

        b3 = tk.Button(popup_cypress, text="Browse for DataLog File:", command=lambda: self.ask_user_file(file_show))
        b3.grid(row=1, column=0)

        cypress_filename_string_var = tk.StringVar(popup_cypress, value=self.cypress_parse_filename)
        file_show = tk.Entry(popup_cypress, textvariable=cypress_filename_string_var, width=150)
        file_show.grid(row=1, column=1)

        b4 = tk.Button(popup_cypress, text="Select Output Folder:", command=lambda:
                       self.ask_user_output_folder(folder_show))
        b4.grid(row=2, column=0)

        excel_destination_string_var = tk.StringVar(popup_cypress, value=self.excel_destination)
        folder_show = tk.Entry(popup_cypress, textvariable=excel_destination_string_var, width=150)
        folder_show.grid(row=2, column=1)

        b1 = tk.Button(popup_cypress, text="Select this DataLog File", command=lambda:
                       self.parse_cypress_ble(popup_cypress, cypress_filename_string_var.get(),
                                              excel_destination_string_var.get()))
        b1.grid(row=3, column=0)

        b2 = tk.Button(popup_cypress, text="Cancel", command=popup_cypress.destroy)
        b2.grid(row=3, column=1)

        popup_cypress.mainloop()

    def cysmart_ble_uv_pop_up(self, num):
        popup_cypress = tk.Toplevel()
        popup_cypress.grab_set()
        popup_cypress.geometry("1000x100")
        popup_cypress.wm_title("Cypress BLE Data Upload")
        label = tk.Label(popup_cypress, text="Upload Cypress BLE Data Log")
        label.grid(row=0, column=0)

        b3 = tk.Button(popup_cypress, text="Browse for DataLog File:", command=lambda: self.ask_user_file(file_show))
        b3.grid(row=1, column=0)

        cypress_filename_string_var = tk.StringVar(popup_cypress, value=self.cypress_parse_filename)
        file_show = tk.Entry(popup_cypress, textvariable=cypress_filename_string_var, width=150)
        file_show.grid(row=1, column=1)

        b4 = tk.Button(popup_cypress, text="Select Output Folder:", command=lambda:
                       self.ask_user_output_folder(folder_show))
        b4.grid(row=2, column=0)

        excel_destination_string_var = tk.StringVar(popup_cypress, value=self.excel_destination)
        folder_show = tk.Entry(popup_cypress, textvariable=excel_destination_string_var, width=150)
        folder_show.grid(row=2, column=1)

        b1 = tk.Button(popup_cypress, text="Select this DataLog File", command=lambda:
                       self.parse_cysmart_uv_ble(popup_cypress, cypress_filename_string_var.get(), excel_destination_string_var.get(), num))
        b1.grid(row=3, column=0)

        b2 = tk.Button(popup_cypress, text="Cancel", command=popup_cypress.destroy)
        b2.grid(row=3, column=1)

        popup_cypress.mainloop()

    def cysmart_ble_temp_pop_up(self, num):
        popup_cypress = tk.Toplevel()
        popup_cypress.grab_set()
        popup_cypress.geometry("1000x100")
        popup_cypress.wm_title("Cypress BLE Data Upload")
        label = tk.Label(popup_cypress, text="Upload Cypress BLE Data Log")
        label.grid(row=0, column=0)

        b3 = tk.Button(popup_cypress, text="Browse for DataLog File:", command=lambda: self.ask_user_file(file_show))
        b3.grid(row=1, column=0)

        cypress_filename_string_var = tk.StringVar(popup_cypress, value=self.cypress_parse_filename)
        file_show = tk.Entry(popup_cypress, textvariable=cypress_filename_string_var, width=150)
        file_show.grid(row=1, column=1)

        b4 = tk.Button(popup_cypress, text="Select Output Folder:", command=lambda:
            self.ask_user_output_folder(folder_show))
        b4.grid(row=2, column=0)

        excel_destination_string_var = tk.StringVar(popup_cypress, value=self.excel_destination)
        folder_show = tk.Entry(popup_cypress, textvariable=excel_destination_string_var, width=150)
        folder_show.grid(row=2, column=1)

        b1 = tk.Button(popup_cypress, text="Select this Temperature DataLog File", command=lambda:
        self.parse_cysmart_temp_ble(popup_cypress, cypress_filename_string_var.get(),
                               excel_destination_string_var.get(), num))
        b1.grid(row=3, column=0)

        b2 = tk.Button(popup_cypress, text="Cancel", command=popup_cypress.destroy)
        b2.grid(row=3, column=1)

        popup_cypress.mainloop()

    def parse_cypress_ble(self, popup_window, source_file_name, destination_file_name):
        popup_window.destroy()
        cypress_datalog_file = str(source_file_name)
        cypress_excel_file = str(destination_file_name)
        cypress_datalog = open(cypress_datalog_file, 'r')
        cypress_datalog_lines = cypress_datalog.readlines()
        cypress_datalog.close()

        tmp116_data = []  # Temperature Data in Celcius from TMP116
        time_stamp = []  # Time Stamp
        seconds = []  # Time Since Data has been Captured
        counter = 0
        time_str_date = 0
        notification = "[Health Thermometer Service|Health Thermometer Measurement] Notification received with value"
        for i in range(len(cypress_datalog_lines)):
            str_temp = ''.join(cypress_datalog_lines.__getitem__(i))

            if notification in str_temp:
                str1_parts = str_temp.split(" , ")
                time_str = ''.join(str1_parts.__getitem__(0))
                temp_str = ''.join(str1_parts.__getitem__(2))
                time_str_temp = time_str[1:-1]
                time_str_date, time_str_clock = time_str_temp.split("|")
                temp_hex = ''.join(temp_str.split(" ").__getitem__(3)) + ''.join(temp_str.split(" ").__getitem__(2))
                temp_dec = int(temp_hex, 16)
                temp_celcius = temp_dec * 7812.5 / 1000000
                tmp116_data.append(temp_celcius)
                time_stamp.append(time_str_temp)
                seconds.append(counter)
                counter = counter + 1

        dest_filename = cypress_excel_file + '/' + time_str_date + '.xlsx'
        png_filename = cypress_excel_file + '/' + time_str_date + '.png'
        wb = Workbook()
        ws = wb.active
        ws.title = "Sample #1"
        ws.cell(row=1, column=1, value="Time Stamp")
        ws.cell(row=1, column=2, value="Seconds")
        ws.cell(row=1, column=3, value="TMP116 (C)")
        wb.save(filename=dest_filename)
        for i in range(1, len(seconds)):
            ws.cell(row=i + 1, column=1, value=time_stamp[i - 1])
            ws.cell(row=i + 1, column=2, value=seconds[i - 1])
            ws.cell(row=i + 1, column=3, value=tmp116_data[i - 1])

        wb.save(filename=dest_filename)

        plt.plot(seconds, tmp116_data)
        plt.ylabel('Temperature (C)')
        plt.xlabel('Time (s)')
        plt.title('TMP116 Temperature')
        plt.savefig(png_filename)
        plt.show()

    def parse_cysmart_uv_ble(self, popup_window, source_file_name, destination_file_name, num):
        popup_window.destroy()
        cypress_datalog_file = str(source_file_name)
        cypress_excel_file = str(destination_file_name)
        cypress_datalog = open(cypress_datalog_file, 'r')
        cypress_datalog_lines = cypress_datalog.readlines()
        cypress_datalog.close()

        uv_data = []  # UV Data in W / cm^2
        time_stamp = []  # Time Stamp
        seconds = []  # Time Since Data has been Captured
        counter = 0
        time_str_date = str(self.day)
        notification = "2E:"
        for i in range(len(cypress_datalog_lines)):
            str_temp = ''.join(cypress_datalog_lines.__getitem__(i))
            if notification in str_temp:
                str1_parts = str_temp.split(":")
                time_hour_temp_str = ''.join(str1_parts.__getitem__(0))
                junk, time_hour_str = time_hour_temp_str.split("[")
                time_min_str = ''.join(str1_parts.__getitem__(1))
                time_sec_str = ''.join(str1_parts.__getitem__(2))
                time_str_temp = time_hour_str + ":" + time_min_str + ":" + time_sec_str
                junk, uv_hex1_str = ''.join(str1_parts.__getitem__(5)).split("[")
                uv_hex2_str = ''.join(str1_parts.__getitem__(6))
                uv_hex3_str = ''.join(str1_parts.__getitem__(7))
                uv_hex4_str = ''.join(str1_parts.__getitem__(8))
                uv_hex5_str, junk = ''.join(str1_parts.__getitem__(9)).split("]")
                uv_str = bytearray.fromhex(uv_hex1_str).decode() + bytearray.fromhex(uv_hex2_str).decode() + bytearray.fromhex(uv_hex3_str).decode() + bytearray.fromhex(
                    uv_hex4_str).decode() + bytearray.fromhex(uv_hex5_str).decode()
                uv_float = float(uv_str)
                uv_data.append(uv_float)
                time_stamp.append(time_str_temp)
                seconds.append(counter)
                counter = counter + 2

        dest_filename = cypress_excel_file + '/' + time_str_date + '.xlsx'
        png_filename = cypress_excel_file + '/' + time_str_date + '.png'
        wb = Workbook()
        ws = wb.active
        ws.title = "Sample #1"
        ws.cell(row=1, column=1, value="Time Stamp")
        ws.cell(row=1, column=2, value="Seconds")
        ws.cell(row=1, column=3, value="UV Power Density (mW/cm^2)")
        wb.save(filename=dest_filename)
        for i in range(1, len(seconds)):
            ws.cell(row=i + 1, column=1, value=time_stamp[i - 1])
            ws.cell(row=i + 1, column=2, value=seconds[i - 1])
            ws.cell(row=i + 1, column=3, value=uv_data[i - 1])

        wb.save(filename=dest_filename)

        plt.plot(seconds, uv_data)
        plt.ylabel('UV Power Density')
        plt.xlabel('Time (s)')
        plt.title('UV Power Density of BLE Sensor Patch')
        plt.savefig(png_filename)
        plt.show()

        if num == 1:
            self.uv_data1 = uv_data
        if num == 2:
            self.uv_data2 = uv_data

    def parse_cysmart_temp_ble(self, popup_window, source_file_name, destination_file_name, num):
        popup_window.destroy()
        cypress_datalog_file = str(source_file_name)
        cypress_excel_file = str(destination_file_name)
        cypress_datalog = open(cypress_datalog_file, 'r')
        cypress_datalog_lines = cypress_datalog.readlines()
        cypress_datalog.close()

        temp_data = []  # Temporary Temperature Data in C
        time_stamp = []  # Time Stamp
        seconds = []  # Time Since Data has been Captured
        start_time = 0
        time_str_date = str(self.day)

        notification = "2E:"
        for i in range(len(cypress_datalog_lines)):
            str_temp = ''.join(cypress_datalog_lines.__getitem__(i))
            if notification in str_temp:
                # Parse Text/Log file to extract temperature values
                str1_parts = str_temp.split(":")
                time_hour_temp_str = ''.join(str1_parts.__getitem__(0))
                junk, time_hour_str = time_hour_temp_str.split("[")
                time_min_str = ''.join(str1_parts.__getitem__(1))
                time_sec_str = ''.join(str1_parts.__getitem__(2))
                time_str_temp = time_hour_str + ":" + time_min_str + ":" + time_sec_str
                junk, temp_hex1_str = ''.join(str1_parts.__getitem__(5)).split("[")
                temp_hex2_str = ''.join(str1_parts.__getitem__(6))
                temp_hex3_str = ''.join(str1_parts.__getitem__(7))
                temp_hex4_str = ''.join(str1_parts.__getitem__(8))
                temp_hex5_str, junk = ''.join(str1_parts.__getitem__(9)).split("]")
                temp_str = bytearray.fromhex(temp_hex1_str).decode() + bytearray.fromhex(temp_hex2_str).decode() + bytearray.fromhex(temp_hex3_str).decode() + \
                    bytearray.fromhex(temp_hex4_str).decode() + bytearray.fromhex(temp_hex5_str).decode()
                temp_float = float(temp_str)    # Convert string number to float number

                if num == 1:
                    self.temp_time_data1.append(datetime(year=self.year, month=self.month, day=self.day, hour=int(time_hour_str), minute=int(time_min_str),
                                                         second=int(time_sec_str)))    # Storing
                    # Time
                    # Objects in an Array
                    self.temp_data1.append(temp_float)  # Storing the float value in an array of temperature values
                    if len(self.temp_time_data1) == 1:
                        seconds.append(0)
                        start_time = self.temp_time_data1[0]
                    else:
                        seconds.append(float((self.temp_time_data1[-1]-start_time).seconds))
                if num == 2:
                    self.temp_time_data2.append(datetime(year=self.year, month=self.month, day=self.day, hour=int(time_hour_str), minute=int(time_min_str),
                                                         second=int(time_sec_str)))    # Storing

                    self.temp_data2.append(temp_float)  # Storing the float value in an array of temperature values
                    if len(self.temp_time_data2) == 1:
                        seconds.append(0)
                        start_time = self.temp_time_data2[0]
                    else:
                        seconds.append(float((self.temp_time_data2[-1] - start_time).seconds))

        if num == 1:
            temp_data = self.temp_data1
            time_stamp = self.temp_time_data1
        if num == 2:
            temp_data = self.temp_data2
            time_stamp = self.temp_time_data2

        dest_filename = cypress_excel_file + '/' + time_str_date + str(num) + '.xlsx'
        self.combined_destination = cypress_excel_file + '/' + time_str_date + 'combined' + '.xlsx'
        png_filename = cypress_excel_file + '/' + time_str_date + str(num) + '.png'
        wb = Workbook()
        ws = wb.active
        ws.title = "Sample #1"
        ws.cell(row=1, column=1, value="Time Stamp")
        ws.cell(row=1, column=2, value="Seconds")
        ws.cell(row=1, column=3, value="Temperature (Celcius)")
        wb.save(filename=dest_filename)
        for i in range(1, len(seconds)):
            ws.cell(row=i + 1, column=1, value=time_stamp[i - 1])
            ws.cell(row=i + 1, column=2, value=seconds[i - 1])
            ws.cell(row=i + 1, column=3, value=temp_data[i - 1])

        wb.save(filename=dest_filename)

        plt.plot(seconds, temp_data)
        plt.ylabel('Temperature')
        plt.xlabel('Time (s)')
        plt.title('Temperature of BLE TMP116')
        plt.savefig(png_filename)
        plt.show()

    def temp_compare(self):
        np_time_data1 = numpy.array(self.temp_time_data1)
        temperature_data1 = pandas.Series(numpy.array(self.temp_data1), index=np_time_data1).resample('1S').mean().interpolate(method='linear')

        np_time_data2 = numpy.array(self.temp_time_data2)
        temperature_data2 = pandas.Series(numpy.array(self.temp_data2), index=np_time_data2).resample('1S').mean().interpolate(method='linear')

        # print(temperature_data1)
        # print(temperature_data2)
        # print("sup")
        # print(temperature_data1.index[0])
        # print(temperature_data2.index[0])
        diff_start = temperature_data2.index[0]-temperature_data1.index[0]
        if diff_start.days < 0:
            first_start_array = temperature_data2
            second_start_array = temperature_data1
        else:
            first_start_array = temperature_data1
            second_start_array = temperature_data2

        first_array_index = 0
        for x in range(len(first_start_array)):
            diff_sec = (first_start_array.index[x] - second_start_array.index[0]).seconds
            if diff_sec == 0:
                first_array_index = x
        first_start_array = first_start_array[first_array_index:]
        # print("starting index")
        # print(first_start_array.index[0])
        # print(second_start_array.index[0])
        #
        # print("Finding Endpoint")
        # print(first_start_array.index[-1])
        # print(second_start_array.index[-1])
        diff_end = first_start_array.index[-1] - second_start_array.index[-1]
        if diff_end.days < 0:
            first_end_array = second_start_array
            second_end_array = first_start_array
        else:
            first_end_array = first_start_array
            second_end_array = second_start_array

        last_array_index = 0
        for x in range(1,len(first_end_array)):
            diff_sec = (first_end_array.index[-x] - second_start_array.index[-1]).seconds
            if diff_sec == 0:
                last_array_index = x-1
        first_end_array = first_end_array[:-last_array_index]

        # print("Endpoint Index")
        # print(first_end_array.index[-1])
        # print(second_end_array.index[-1])
        if first_start_array.index[0] == second_start_array.index[0]:
            print("Start match up")
        if first_end_array.index[-1] == second_end_array.index[-1]:
            print("Endpoints match up")

        temp_diff = first_end_array.values - second_end_array.values
        # print(first_end_array.values[0])
        # print(second_end_array.values[0])
        # print(temp_diff[0])
        std_value = statistics.pstdev(temp_diff)
        avg_value = statistics.mean(temp_diff)
        # print(std_value)
        seconds = numpy.arange(len(first_end_array.values))

        plt.figure(3)
        plt.plot(seconds, first_end_array.values, label="BLE Patch", linestyle='-')
        plt.plot(seconds, second_end_array.values, label="PCB", linestyle='-')
        plt.ylabel('Temperature')
        plt.xlabel('Time (s)')
        plt.title('Temperature of BLE TMP116 & PCB TMP116, std = ' + str(round(std_value,4)) + ', AVG diff=' + str(round(avg_value,4)))
        plt.legend()
        plt.show()
        print(self.combined_destination)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sample Test"
        ws2.cell(row=1, column=1, value="Time Stamp")
        ws2.cell(row=1, column=2, value="BLE Patch (Celcius)")
        ws2.cell(row=1, column=3, value="PCB (Celcius)")
        wb2.save(filename=self.combined_destination)
        for i in range(1, len(seconds)):
            ws2.cell(row=i + 1, column=1, value=seconds[i - 1])
            ws2.cell(row=i + 1, column=2, value=first_end_array.values[i - 1])
            ws2.cell(row=i + 1, column=3, value=second_end_array.values[i - 1])

        wb2.save(filename=self.combined_destination)
        print(self.combined_destination)

def main():
    init_gui()


if __name__ == "__main__":
    main()
