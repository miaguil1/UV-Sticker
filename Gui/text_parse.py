import matplotlib.pyplot as plt
from openpyxl import Workbook
import time

file = open('28-Oct-2017.txt', 'r')
lines = file.readlines()
file.close()

TMP116_Data = []    # Temperature Data in Celcius from TMP116
Time_Stamp = []     # Time Stamp
seconds = [] # Time Since Data has been Captured
counter = 0
notification = "[Health Thermometer Service|Health Thermometer Measurement] Notification received with value"
for i in range(len(lines)):
    str_temp = ''.join(lines.__getitem__(i))

    if notification in str_temp:
        str1_parts = str_temp.split(" , ")
        time_str = ''.join(str1_parts.__getitem__(0))
        temp_str = ''.join(str1_parts.__getitem__(2))
        time_str_temp = time_str[1:-1]
        time_str_date, time_str_clock = time_str_temp.split("|")
        # print(temp_str)
        temp_hex = ''.join(temp_str.split(" ").__getitem__(3)) + ''.join(temp_str.split(" ").__getitem__(2))
        temp_dec = int(temp_hex, 16)
        temp_celcius = temp_dec * 7812.5 / 1000000
        TMP116_Data.append(temp_celcius)
        Time_Stamp.append(time_str_temp)
        seconds.append(counter)
        counter = counter + 1
# print(counter)
dest_filename = 'C:/Users/Intern/Documents/Data/test-temp_1.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "Sample #1"
ws.cell(row=1, column=1, value="Time Stamp")
ws.cell(row=1, column=2, value="Seconds")
ws.cell(row=1, column=3, value="TMP116 (C)")
wb.save(filename = dest_filename)
for i in range(1, len(seconds)):
    ws.cell(row=i+1, column=1, value=Time_Stamp[i-1])
    ws.cell(row=i+1, column=2, value=seconds[i-1])
    ws.cell(row=i+1, column=3, value=TMP116_Data[i-1])

wb.save(filename = dest_filename)

plt.plot(seconds, TMP116_Data)
plt.ylabel('Temperature (C)')
plt.xlabel('Time (s)')
plt.title('TMP116 Temperature')
plt.show()